"""
openpyxl简单使用：https://blog.csdn.net/weixin_46371752/article/details/124681081?utm_medium=distribute.pc_relevant.none-task-blog-2~default~baidujs_baidulandingword~default-9-124681081-blog-125570707.235^v36^pc_relevant_default_base&spm=1001.2101.3001.4242.6&utm_relevant_index=12
安装：pip3 install openpyxl
wb = openpyxl.load_workbook('test.xlsx')
ws = wb['运维'] 或 ws=wb.active
#获取工作表的最大行数
max_row = ws.max_row
#获取工作表的最小行数
min_row = ws.min_row
#获取工作表的最大列数
max_col = ws.max_column
#获取工作表的最小列数
min_col = ws.min_column

ws['A1'].value  #A1等于1列1行
ws.cell(row=i,column=2).value
ws['G2']='新写入的数据'#写入数据方式

cell_range=ws['A1:E5']#获取范围:A1:E5
for rows in cell_range:
    for data in rows:
        print(data.value)#这个for提取的是各个单元格的数据

range_=ws.iter_rows(min_row=1,max_row=1,min_col=1,max_col=5)#指定范围提取，min_row为最小行,max_row为最大行,min_col为最小列,max_col为最大列。
for rows in cell_range:
    for cells in rows:
        print(cells.value)  #按行，依次输出单元格值
"""
import openpyxl
# openpyxl 3.1.2
from openpyxl.cell import MergedCell

def read_row_cell(line_num, start_index, end_index, interval=1):
    """
    带间隔功能的读取一行多个cell
    :param line_num:
    :param start_index:
    :param end_index:
    :param interval:
    :return:
    """
    for index in range(start_index, end_index+1, interval):
        cell = ws.cell(row=line_num,column=index)
        print(cell.value)

def read_column_cell(column_num, start_index, end_index, interval=1):
    """
    带间隔功能的读取一列的多个cell
    :param column_num:
    :param start_index:
    :param end_index:
    :param interval:
    :return:
    """
    for index in range(start_index, end_index + 1, interval):
        cell = ws.cell(row=index, column=column_num)
        print(cell.value)

def is_in_merge_cell(row, column):
    """
    判断一个cell是否属于mergecell
    :param row:
    :param column:
    :return: bool
    """
    cell = ws.cell(row, column)
    for merged_cell in ws.merged_cells.ranges:
        # cell.coordinate like A1
        # ws.merged_cells.ranges：表中mergecell的集合set
        # 对于mergecell只有最左上角位置的cell有值，其余为None
        if (cell.coordinate in merged_cell):
            return True
    return False

def get_cell_value(row, column):
    """
    获取某个cell/mergecell的value
    :param row:
    :param column:
    :return:
    """
    cell = ws.cell(row=row, column=column)
    if isinstance(cell, MergedCell):  # 判断该单元格是否为合并单元格
        for merged_cell in ws.merged_cells.ranges:
            if (cell.coordinate in merged_cell):
                # cell有col和row，mergecell有min/max_row，min/max_col
                cell = ws.cell(row=merged_cell.min_row, column=merged_cell.min_col)
                return cell.value
    else:
        return cell.value

def traverse_row_cell(row):
    """
    遍历一行所有cell的value(不是通过索引遍历的，索引遍历在上面案例中)。最好不要有mergecell，会拿到None。
    :param row:
    :return:
    """
    cell_range = ws.iter_rows(min_row=row, max_row=row, min_col=1,
                           max_col=ws.max_column)  # 指定范围提取，min_row为最小行,max_row为最大行,min_col为最小列,max_col为最大列。
    for rows in cell_range:
        for cell in rows:
            print(cell.value)  # 按行，依次输出单元格值


if __name__ == '__main__':
    wb = openpyxl.load_workbook('test.xlsx')
    ws = wb.active
    print(ws.max_column)
    print(ws.max_row)
    read_row_cell(5,1,3,2)
    read_column_cell(1,1,5,2)
    print(is_in_merge_cell(3,5))
    traverse_row_cell(5)

    print(ws.merged_cells.ranges)  # {<MergedCellRange F5:G5>, <MergedCellRange D1:E2>, <MergedCellRange C3:C4>, <MergedCellRange F1:G2>}
    print(ws.merged_cells)  # C3:C4 D1:E2 F1:G2 F5:G5
    for merged_cell in ws.merged_cells.ranges:
        print(merged_cell)  # C3:C4 类型MergedCellRange

    print(get_cell_value(1,5))
